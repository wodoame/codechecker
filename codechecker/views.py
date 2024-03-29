import pandas as pd
from django.shortcuts import render, redirect
from .forms import CheckUpdateForm
from .models import Record
from django.http import HttpResponse
from django.views import View
from django.views.generic import TemplateView
from django.forms.models import modelform_factory
from .models import Event, AppData
import csv
from django.contrib import messages
from datetime import datetime
from openpyxl import load_workbook
import os
from collections import deque, Counter

def populate_model_from_spreadsheet(data_file):
    # Load the data from the spreadsheet
    data = pd.read_excel(data_file)

    # Create a set to keep track of processed codes
    processed_codes = set()

    # Iterate through the data to populate the model
    for index, row in data.iterrows():
        code = row['code']

        # Check if this code has already been processed
        if code in processed_codes:
            continue

        msisdn = row['msisdn']
        first_name = row['first_name']
        last_name = row['last_name']

        try:
            record = Record.objects.get(code=code)
            # Update the fields if they are different in the spreadsheet
            if record.msisdn != msisdn or record.first_name != first_name or record.last_name != last_name:
                record.msisdn = msisdn
                record.first_name = first_name
                record.last_name = last_name
                record.save()
        except Record.DoesNotExist:
            # If the record doesn't exist in the model, create it
            record = Record.objects.create(
                code=code,
                msisdn=msisdn,
                first_name=first_name,
                last_name=last_name,
                attended='no',
                no_of_uses=0,
                number='No number given'
            )

        # Mark the code as processed
        processed_codes.add(code)

    # Handle deletions (if a record in the model doesn't exist in the spreadsheet)
    codes_in_spreadsheet = set(data['code'])
    codes_in_model = set(Record.objects.values_list('code', flat=True))
    codes_to_delete = codes_in_model - codes_in_spreadsheet
    Record.objects.filter(code__in=codes_to_delete).delete()


def check(request):
    data_file = 'data/your_data.xlsx'  # Update with your spreadsheet file path
    # Call the function to populate the model (only on the initial run)
    if not Record.objects.exists():
        populate_model_from_spreadsheet(data_file)

    if request.method == 'POST':
        form = CheckUpdateForm(request.POST)

        if form.is_valid():
            code_to_check = form.cleaned_data['code']
            number = form.cleaned_data['number']
            
            try:
                record = Record.objects.get(code=code_to_check)
                if record.attended == 'no' and record.no_of_uses == 0:
                    # First use of the code
                    record.attended = 'yes'
                    record.no_of_uses = 1
                    record.number = number if number else 'No number given'
                    record.save()
                    message = f'Code used for the first time. Number appended'
                    message_class = 'success'
                elif record.attended == 'yes' and record.no_of_uses == 1:
                    # Second use of the code
                    record.no_of_uses = 2
                    record.save()
                    message = 'Code used for the second time. Number remains the same.'
                    message_class = 'success'
                else:
                    # Code has been used more than twice
                    message = 'Code cannot be used again.'
                    message_class = 'error'
            except Record.DoesNotExist:
                # Code not found in the model
                message = 'Code not found in the model.'
                message_class = 'error'
        else:
            message = 'Invalid input. Please try again.'
            message_class = 'error'
    else:
        form = CheckUpdateForm()
        message = ''
        message_class = ''
    
    new_form = CheckUpdateForm()

    return render(request, 'check.html', {'form': new_form, 'message': message, 'message_class': message_class})


def export_records(request):
    # Query the records from the model
    records = Record.objects.all()

    # Create a DataFrame from the records
    data = {
        'MSISDN': [record.msisdn for record in records],
        'First Name': [record.first_name for record in records],
        'Last Name': [record.last_name for record in records],
        'Code': [record.code for record in records],
        'Attended': [record.attended for record in records],
        'Number': [record.number for record in records],
    }
    df = pd.DataFrame(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="records.xlsx"'
    df.to_excel(response, index=False, engine='openpyxl')

    return response


# Extension

class SetupEvent(View):
    form_class = modelform_factory(model=Event, fields='__all__')
    def get(self, request): 
        return render(request, 'setup.html')
    
    def process_file(self, db, is_excel_file):
        filepath = db.dataset.path 
        newpath = filepath
        if is_excel_file:
            newpath = filepath.removesuffix('.xlsx') + '.csv'
            wb = load_workbook(filepath)
            sheet = wb.active
            data = sheet.iter_rows(values_only=True)
            headers = next(data)
            headers = [key.lower() for key in headers]  # making the header lowercase for uniformity
            with open(newpath, 'w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)
                csv_writer.writerow(headers)
                csv_writer.writerows(data)
                db.dataset.name = newpath
                db.save()
                os.remove(filepath)
        else: 
            with open(newpath, 'r', newline='') as uploaded_file:
                reader = csv.reader(uploaded_file)
                headers = next(reader)
                headers = [key.lower() for key in headers]  # making the header lowercase for uniformity
                with open(newpath, 'w', newline='') as csv_file: 
                    writer = csv.writer(csv_file)
                    writer.writerow(headers)
                    writer.writerows(reader)
   
            
              
    def post(self, request):
        EXTENSIONS = ['.csv', '.xlsx']
        form = self.form_class(request.POST, request.FILES)
        filename:str = request.FILES.get('dataset').name 
        if not any([extension in filename for extension in EXTENSIONS]): 
           messages.success(request, 'File should be a CSV or EXCEL file')
        else:
             if form.is_valid():
                messages.success(request, f"{request.POST.get('event_title')} created successfully") 
                db = form.save()
                is_excel_file = filename.endswith(EXTENSIONS[1])
                self.process_file(db, is_excel_file)
             else: 
                print(form.errors)
        return redirect('dashboard')
    
class ViewEvent(View):
    template_name='event.html'
    def blank(self, request, context={}):
        # request._messages = messages.storage.default_storage(request)
        # manually clear messages (I didn't need it)
        return render(request, self.template_name, context)
    
    def get(self, request, slug):
        db = Event.objects.get(slug=slug)
        attendanceRecords = db.attendance_records.filter(date_created=datetime.now().date()).first()
        presentToday = attendanceRecords.present if attendanceRecords else []
        headers = []
        if presentToday: 
            headers = presentToday[0]
        if request.user.is_authenticated:
            app_data, created = AppData.objects.get_or_create(user=request.user)
            recently_opened = app_data.recently_opened
            recently_opened = deque(recently_opened)
            recently_opened.appendleft(db.slug)
            app_data.recently_opened = list(Counter(recently_opened).keys()) # using Counter.values to get a unique but ordered list
            app_data.save()
        
        return render(request, self.template_name, {'db':db, 'presentToday': presentToday, 'headers':headers})
    
    def post(self, request, slug): 
        validation_credentials = request.POST.get('validation_field')
        db =  Event.objects.get(slug=slug)
        attendanceRecordsToday, created = db.attendance_records.get_or_create(date_created=datetime.now().date())
        # If person is already checked in then we can exit the function quickly
        # Otherwise we check every record until we get a match 
        # If we don't get a match then there is no such person
        if attendanceRecordsToday.checked_in.get(validation_credentials):
            messages.info(request, 'Already checked in') 
            return self.blank(request, {'checked_in':True, 'db':db, 'tag': 'success'})
        
        with open(db.dataset.path, 'r', newline='') as dataset: 
            reader = csv.DictReader(dataset)
            for data in reader:
                if validation_credentials == data.get(db.validation_field):
                    attendanceRecordsToday.checked_in.update({validation_credentials:True})
                    attendanceRecordsToday.present.append(data)
                    attendanceRecordsToday.save()
                    messages.success(request, 'Checked in successfully')
                    return self.blank(request, {'checked_in':True, 'db':db, 'tag': 'success'})
        messages.error(request, 'No such credentials')
        return self.blank(request, {'checked_in':True, 'db':db, 'tag':'danger'})
        
class Dashboard(View):
    template_name = 'dashboard.html'
    def get(self, request):
        # print(request.META.get('HTTP_COOKIE'))
        databases = Event.objects.all()
        if request.user.is_authenticated: 
            app_data, created = AppData.objects.get_or_create(user=request.user)
            recently_opened = [] # This is going to contain actual objects that will be sent to the template
            for item in app_data.recently_opened: # app_data.recently_opened contain slugs of events that are recently opened
                recently_opened.append(Event.objects.get(slug=item))
                
        most_attended = self.most_attended(Event.objects.all())
        context = {'databases':databases, 'recently_opened':recently_opened, 'most_attended':most_attended}
        return render(request, self.template_name, context)
    
    def most_attended(self, events):
        averages = []
        for event in events: 
            total = 0
            count = 0
            average = 0
            for record in event.attendance_records.all():
                total += len(record.present)
                count += 1
            if count > 0: # prevent division by zero
                average = total / count
            averages.append({'event': event, 'average_attendance': average})
        
        # sorting 
        k = len(averages)
        for i in range(k): 
            for j in range(k - 1): 
                if averages[j].get('average_attendance') < averages[j + 1].get('average_attendance'):
                    averages[j], averages[j + 1] = averages[j + 1], averages[j]
        return averages
                
        
            
        
        
            
            
                
                
            


class AddPerson(View):
    template_name = 'add-person.html'
    def get(self, request, slug):
        db = Event.objects.get(slug=slug)
        with open(db.dataset.path, 'r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            fields = next(csv_reader) # Getting the names of the fields
        return render(request, self.template_name, {'db': db, 'fields': fields})
    
    def post(self, request, slug):
        return self.get(request, slug)
        
        
        
        
    
            
        
    
    
    


